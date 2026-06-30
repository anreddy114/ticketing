"""
Backend regression + new-feature tests for ticketing system iteration 2:
- /api/settings (GET/PATCH)
- /api/agents (list with online status)
- /api/agents/heartbeat
- /api/agents/presence (with all 4 offline strategies)
- /api/sip/incoming-call (with secret guard)
- /api/ivr/event, /api/ivr/agent-availability
- /api/reports/tickets.xlsx (admin + agent + filters)
- Regression: /api/auth/login, /api/tickets list, /api/issue-types
"""
import os
import io
import time
import pytest
import requests
from pathlib import Path

# Internal URL per testing instructions (other_misc_info)
BASE_URL = "http://localhost:8001"

ADMIN_EMAIL = "admin@ticketing.com"
ADMIN_PASS = "admin123"
AGENT_EMAIL = "agent@ticketing.com"
AGENT_PASS = "agent123"


def _read_env_secret() -> str:
    env_path = Path("/app/backend/.env")
    for line in env_path.read_text().splitlines():
        if line.startswith("SIP_WEBHOOK_SECRET"):
            return line.split("=", 1)[1].strip().strip('"').strip("'")
    return ""


SIP_SECRET = _read_env_secret()


# -------------------- fixtures --------------------

@pytest.fixture(scope="session")
def admin_token():
    r = requests.post(f"{BASE_URL}/api/auth/login",
                      json={"email": ADMIN_EMAIL, "password": ADMIN_PASS}, timeout=15)
    assert r.status_code == 200, f"Admin login failed: {r.status_code} {r.text}"
    return r.json()["token"]


@pytest.fixture(scope="session")
def agent_token():
    r = requests.post(f"{BASE_URL}/api/auth/login",
                      json={"email": AGENT_EMAIL, "password": AGENT_PASS}, timeout=15)
    if r.status_code != 200:
        # Register the agent via admin
        a = requests.post(f"{BASE_URL}/api/auth/login",
                          json={"email": ADMIN_EMAIL, "password": ADMIN_PASS}, timeout=15)
        atoken = a.json()["token"]
        requests.post(
            f"{BASE_URL}/api/auth/register",
            headers={"Authorization": f"Bearer {atoken}"},
            json={"email": AGENT_EMAIL, "password": AGENT_PASS,
                  "name": "Test Agent", "role": "agent"},
            timeout=15,
        )
        r = requests.post(f"{BASE_URL}/api/auth/login",
                          json={"email": AGENT_EMAIL, "password": AGENT_PASS}, timeout=15)
    assert r.status_code == 200, f"Agent login failed: {r.status_code} {r.text}"
    return r.json()["token"]


def _auth(token: str) -> dict:
    return {"Authorization": f"Bearer {token}"}


@pytest.fixture(scope="session")
def admin_user(admin_token):
    r = requests.get(f"{BASE_URL}/api/auth/me", headers=_auth(admin_token), timeout=10)
    return r.json()


@pytest.fixture(scope="session")
def agent_user(agent_token):
    r = requests.get(f"{BASE_URL}/api/auth/me", headers=_auth(agent_token), timeout=10)
    return r.json()


# -------------------- regression: auth --------------------

class TestAuthRegression:
    def test_admin_login(self, admin_token):
        assert isinstance(admin_token, str) and len(admin_token) > 10

    def test_agent_login(self, agent_token):
        assert isinstance(agent_token, str) and len(agent_token) > 10

    def test_me_admin(self, admin_user):
        assert admin_user["email"] == ADMIN_EMAIL
        assert admin_user["role"] == "admin"


# -------------------- /api/settings --------------------

class TestSettings:
    def test_get_settings_requires_auth(self):
        r = requests.get(f"{BASE_URL}/api/settings", timeout=10)
        assert r.status_code in (401, 403), f"Expected unauth, got {r.status_code}"

    def test_get_settings_returns_expected_keys(self, admin_token):
        r = requests.get(f"{BASE_URL}/api/settings", headers=_auth(admin_token), timeout=10)
        assert r.status_code == 200, r.text
        d = r.json()
        for k in ("offline_strategy", "fallback_assignee_id",
                  "sip_webhook_url", "ivr_webhook_url", "heartbeat_timeout_sec"):
            assert k in d, f"Missing key {k}: {d}"
        assert d["offline_strategy"] in ("stay", "round_robin", "fallback", "manual_transfer")
        assert d["heartbeat_timeout_sec"] == 120

    def test_patch_settings_non_admin_403(self, agent_token):
        r = requests.patch(f"{BASE_URL}/api/settings",
                           headers=_auth(agent_token),
                           json={"offline_strategy": "stay"}, timeout=10)
        assert r.status_code == 403, f"Got {r.status_code}: {r.text}"

    def test_patch_settings_admin_updates(self, admin_token):
        r = requests.patch(f"{BASE_URL}/api/settings",
                           headers=_auth(admin_token),
                           json={"offline_strategy": "round_robin"}, timeout=10)
        assert r.status_code == 200, r.text
        assert r.json()["offline_strategy"] == "round_robin"


# -------------------- /api/agents + heartbeat + presence --------------------

class TestAgents:
    def test_list_agents(self, admin_token):
        r = requests.get(f"{BASE_URL}/api/agents", headers=_auth(admin_token), timeout=10)
        assert r.status_code == 200, r.text
        data = r.json()
        assert isinstance(data, list) and len(data) >= 1
        for u in data:
            assert "online" in u and isinstance(u["online"], bool)
            assert "password_hash" not in u
            assert "_id" not in u

    def test_heartbeat(self, agent_token):
        r = requests.post(f"{BASE_URL}/api/agents/heartbeat",
                          headers=_auth(agent_token), timeout=10)
        assert r.status_code == 200, r.text
        assert r.json().get("ok") is True

    def test_presence_online_returns_zero_reassigned(self, agent_token):
        r = requests.post(f"{BASE_URL}/api/agents/presence",
                          headers=_auth(agent_token),
                          json={"online": True}, timeout=10)
        assert r.status_code == 200, r.text
        body = r.json()
        assert body["online"] is True
        assert body["tickets_reassigned"] == 0


# -------------------- offline strategy: round_robin redistribution --------------------

class TestRoundRobinRedistribution:
    """Verify open tickets assigned to admin get moved to online agent when admin goes offline."""

    def test_round_robin_flow(self, admin_token, agent_token, admin_user, agent_user):
        # 1. Ensure offline_strategy = round_robin
        r = requests.patch(f"{BASE_URL}/api/settings",
                          headers=_auth(admin_token),
                          json={"offline_strategy": "round_robin"}, timeout=10)
        assert r.status_code == 200

        # 2. Agent goes online + heartbeat
        r = requests.post(f"{BASE_URL}/api/agents/presence",
                          headers=_auth(agent_token),
                          json={"online": True}, timeout=10)
        assert r.status_code == 200
        requests.post(f"{BASE_URL}/api/agents/heartbeat",
                      headers=_auth(agent_token), timeout=10)

        # 3. Admin goes online
        r = requests.post(f"{BASE_URL}/api/agents/presence",
                          headers=_auth(admin_token),
                          json={"online": True}, timeout=10)
        assert r.status_code == 200

        # 4. Get an issue_type id
        it = requests.get(f"{BASE_URL}/api/issue-types",
                          headers=_auth(admin_token), timeout=10).json()
        assert len(it) > 0
        issue_id = it[0]["id"]

        # 5. Create a ticket assigned to admin (source=self)
        ticket_payload = {
            "source": "self",
            "issue_type_id": issue_id,
            "title": "TEST_RR redistribution",
            "description": "test",
            "priority": "medium",
            "assigned_to": admin_user["id"],
        }
        r = requests.post(f"{BASE_URL}/api/tickets",
                          headers=_auth(admin_token),
                          json=ticket_payload, timeout=10)
        assert r.status_code in (200, 201), r.text
        tkt = r.json()
        tid = tkt["id"]
        assert tkt["assigned_to"] == admin_user["id"]

        # 6. Admin goes offline → should redistribute to agent (the only other online)
        r = requests.post(f"{BASE_URL}/api/agents/presence",
                          headers=_auth(admin_token),
                          json={"online": False}, timeout=15)
        assert r.status_code == 200, r.text
        body = r.json()
        assert body["online"] is False
        assert body["tickets_reassigned"] >= 1, f"Expected >=1 reassigned, got {body}"

        # 7. Verify ticket is now assigned to agent
        r = requests.get(f"{BASE_URL}/api/tickets/{tid}",
                         headers=_auth(admin_token), timeout=10)
        assert r.status_code == 200
        t = r.json()["ticket"]
        assert t["assigned_to"] == agent_user["id"], \
            f"Ticket not reassigned to agent: {t['assigned_to']} vs {agent_user['id']}"


# -------------------- offline strategy: manual_transfer --------------------

class TestManualTransferStrategy:
    def test_manual_transfer_requires_transfer_to(self, admin_token, agent_token):
        # set strategy to manual_transfer
        requests.patch(f"{BASE_URL}/api/settings",
                       headers=_auth(admin_token),
                       json={"offline_strategy": "manual_transfer"}, timeout=10)

        # agent must be currently online for redistribution code path to run
        requests.post(f"{BASE_URL}/api/agents/presence",
                      headers=_auth(agent_token),
                      json={"online": True}, timeout=10)

        r = requests.post(f"{BASE_URL}/api/agents/presence",
                          headers=_auth(agent_token),
                          json={"online": False}, timeout=10)
        assert r.status_code == 400, f"Expected 400, got {r.status_code}: {r.text}"

    def test_manual_transfer_with_valid_target(self, admin_token, agent_token, admin_user, agent_user):
        # Set strategy
        requests.patch(f"{BASE_URL}/api/settings",
                       headers=_auth(admin_token),
                       json={"offline_strategy": "manual_transfer"}, timeout=10)

        # Both online
        requests.post(f"{BASE_URL}/api/agents/presence",
                      headers=_auth(agent_token),
                      json={"online": True}, timeout=10)
        requests.post(f"{BASE_URL}/api/agents/presence",
                      headers=_auth(admin_token),
                      json={"online": True}, timeout=10)

        # Create a ticket assigned to agent
        it = requests.get(f"{BASE_URL}/api/issue-types",
                          headers=_auth(admin_token), timeout=10).json()
        r = requests.post(f"{BASE_URL}/api/tickets",
                          headers=_auth(admin_token),
                          json={
                              "source": "self",
                              "issue_type_id": it[0]["id"],
                              "title": "TEST_manual_transfer",
                              "description": "x",
                              "priority": "low",
                              "assigned_to": agent_user["id"],
                          }, timeout=10)
        assert r.status_code in (200, 201), r.text
        tid = r.json()["id"]

        # Agent goes offline with transfer_to = admin
        r = requests.post(f"{BASE_URL}/api/agents/presence",
                          headers=_auth(agent_token),
                          json={"online": False, "transfer_to": admin_user["id"]}, timeout=15)
        assert r.status_code == 200, r.text
        assert r.json()["tickets_reassigned"] >= 1

        # Verify
        r = requests.get(f"{BASE_URL}/api/tickets/{tid}",
                         headers=_auth(admin_token), timeout=10)
        assert r.json()["ticket"]["assigned_to"] == admin_user["id"]


# -------------------- offline strategy: fallback --------------------

class TestFallbackStrategy:
    def test_fallback(self, admin_token, agent_token, admin_user, agent_user):
        # Set fallback_assignee_id = admin, strategy=fallback
        r = requests.patch(f"{BASE_URL}/api/settings",
                           headers=_auth(admin_token),
                           json={"offline_strategy": "fallback",
                                 "fallback_assignee_id": admin_user["id"]}, timeout=10)
        assert r.status_code == 200, r.text

        # Both online
        requests.post(f"{BASE_URL}/api/agents/presence",
                      headers=_auth(agent_token),
                      json={"online": True}, timeout=10)
        requests.post(f"{BASE_URL}/api/agents/presence",
                      headers=_auth(admin_token),
                      json={"online": True}, timeout=10)

        # Create ticket assigned to agent
        it = requests.get(f"{BASE_URL}/api/issue-types",
                          headers=_auth(admin_token), timeout=10).json()
        r = requests.post(f"{BASE_URL}/api/tickets",
                          headers=_auth(admin_token),
                          json={
                              "source": "self",
                              "issue_type_id": it[0]["id"],
                              "title": "TEST_fallback",
                              "description": "x",
                              "priority": "low",
                              "assigned_to": agent_user["id"],
                          }, timeout=10)
        tid = r.json()["id"]

        # Agent goes offline (no transfer_to needed)
        r = requests.post(f"{BASE_URL}/api/agents/presence",
                          headers=_auth(agent_token),
                          json={"online": False}, timeout=15)
        assert r.status_code == 200, r.text
        assert r.json()["tickets_reassigned"] >= 1

        # Verify moved to admin (fallback)
        r = requests.get(f"{BASE_URL}/api/tickets/{tid}",
                         headers=_auth(admin_token), timeout=10)
        assert r.json()["ticket"]["assigned_to"] == admin_user["id"]


# -------------------- offline strategy: stay --------------------

class TestStayStrategy:
    def test_stay(self, admin_token, agent_token, agent_user):
        requests.patch(f"{BASE_URL}/api/settings",
                       headers=_auth(admin_token),
                       json={"offline_strategy": "stay"}, timeout=10)

        # agent online
        requests.post(f"{BASE_URL}/api/agents/presence",
                      headers=_auth(agent_token),
                      json={"online": True}, timeout=10)

        # ticket to agent
        it = requests.get(f"{BASE_URL}/api/issue-types",
                          headers=_auth(admin_token), timeout=10).json()
        r = requests.post(f"{BASE_URL}/api/tickets",
                          headers=_auth(admin_token),
                          json={
                              "source": "self",
                              "issue_type_id": it[0]["id"],
                              "title": "TEST_stay",
                              "description": "x",
                              "priority": "low",
                              "assigned_to": agent_user["id"],
                          }, timeout=10)
        tid = r.json()["id"]

        # offline
        r = requests.post(f"{BASE_URL}/api/agents/presence",
                          headers=_auth(agent_token),
                          json={"online": False}, timeout=15)
        assert r.status_code == 200
        assert r.json()["tickets_reassigned"] == 0

        # Ticket still on agent
        r = requests.get(f"{BASE_URL}/api/tickets/{tid}",
                         headers=_auth(admin_token), timeout=10)
        assert r.json()["ticket"]["assigned_to"] == agent_user["id"]


# -------------------- SIP webhook --------------------

class TestSipWebhook:
    def test_sip_missing_secret_returns_401(self):
        r = requests.post(f"{BASE_URL}/api/sip/incoming-call",
                          json={"caller_mobile": "9999900001"}, timeout=15)
        assert r.status_code == 401, f"Expected 401, got {r.status_code}: {r.text}"

    def test_sip_creates_ticket(self, admin_token, agent_token):
        # Bring at least one agent online so round-robin has a candidate
        requests.patch(f"{BASE_URL}/api/settings",
                       headers=_auth(admin_token),
                       json={"offline_strategy": "round_robin"}, timeout=10)
        requests.post(f"{BASE_URL}/api/agents/presence",
                      headers=_auth(agent_token),
                      json={"online": True}, timeout=10)
        requests.post(f"{BASE_URL}/api/agents/heartbeat",
                      headers=_auth(agent_token), timeout=10)

        assert SIP_SECRET, "SIP_WEBHOOK_SECRET not loaded from .env"
        payload = {
            "caller_mobile": "9999900001",
            "call_id": "TEST_CALL_001",
            "did": "+912245678900",
            "agent_busy": True,
            "notes": "queue overflow",
        }
        r = requests.post(f"{BASE_URL}/api/sip/incoming-call",
                          json=payload,
                          headers={"X-Sip-Secret": SIP_SECRET}, timeout=30)
        assert r.status_code in (200, 201), f"{r.status_code}: {r.text}"
        t = r.json()
        assert t.get("ticket_number", "").startswith("TKT-"), t
        assert t.get("origin_channel") == "sip_inbound"
        assert t.get("customer_mobile") in ("9999900001",) or t.get("customer_mobile")
        assert t.get("customer_name"), "customer_name should be populated"
        assert t.get("assigned_to"), "assignee should be set"


# -------------------- IVR webhook --------------------

class TestIvr:
    def test_ivr_event_records(self):
        assert SIP_SECRET
        payload = {
            "caller_mobile": "9999900001",
            "event": "dtmf",
            "payload": {"digits": "1"},
            "call_id": "TEST_IVR_001",
        }
        r = requests.post(f"{BASE_URL}/api/ivr/event",
                          json=payload,
                          headers={"X-Sip-Secret": SIP_SECRET}, timeout=10)
        assert r.status_code == 200, r.text
        body = r.json()
        assert body["ok"] is True
        assert body["event"]["event"] == "dtmf"
        assert body["event"]["caller_mobile"] == "9999900001"
        assert "_id" not in body["event"]

    def test_ivr_agent_availability(self, admin_token, agent_token):
        # Ensure agent online
        requests.post(f"{BASE_URL}/api/agents/presence",
                      headers=_auth(agent_token),
                      json={"online": True}, timeout=10)
        requests.post(f"{BASE_URL}/api/agents/heartbeat",
                      headers=_auth(agent_token), timeout=10)

        # endpoint requires SIP secret
        r = requests.get(f"{BASE_URL}/api/ivr/agent-availability",
                         headers={"X-Sip-Secret": SIP_SECRET}, timeout=10)
        assert r.status_code == 200, r.text
        d = r.json()
        assert "online_count" in d and "available" in d
        assert d["online_count"] >= 1
        assert d["available"] is True


# -------------------- Reports xlsx --------------------

class TestReportsXlsx:
    def test_admin_xlsx_download(self, admin_token):
        r = requests.get(f"{BASE_URL}/api/reports/tickets.xlsx",
                         headers=_auth(admin_token), timeout=30)
        assert r.status_code == 200, r.text[:200]
        ct = r.headers.get("content-type", "")
        assert "spreadsheetml" in ct, f"content-type={ct}"
        assert len(r.content) > 1000, f"file too small: {len(r.content)}"
        assert r.content[:2] == b"PK", "xlsx must start with PK"

    def test_xlsx_with_filters(self, admin_token):
        r = requests.get(
            f"{BASE_URL}/api/reports/tickets.xlsx",
            params={"status": "open",
                    "date_from": "2024-01-01T00:00:00Z",
                    "date_to": "2030-12-31T23:59:59Z"},
            headers=_auth(admin_token), timeout=30)
        assert r.status_code == 200
        assert r.content[:2] == b"PK"

    def test_agent_xlsx_scoped(self, agent_token, agent_user):
        r = requests.get(f"{BASE_URL}/api/reports/tickets.xlsx",
                         headers=_auth(agent_token), timeout=30)
        assert r.status_code == 200, r.text[:200]
        assert r.content[:2] == b"PK"
        # Open the workbook and verify all rows are assigned to agent
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(r.content))
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        # header + at least 0 rows; if there are data rows check assignee column
        # Headers: Ticket #, Title, Status, Priority, Issue Type, Source,
        # Customer Name, Customer Mobile, Package, Expiry, Partner,
        # Assigned To, Created By, Created At, Closed At  -> assigned_to is index 11
        for row in rows[1:]:
            if row[0] is None:
                continue
            assigned_name = row[11]
            assert assigned_name == agent_user["name"], \
                f"Agent xlsx leaked another assignee: {assigned_name}"


# -------------------- Regression: tickets listing --------------------

class TestTicketRegression:
    def test_tickets_list(self, admin_token):
        r = requests.get(f"{BASE_URL}/api/tickets",
                         headers=_auth(admin_token), timeout=10)
        assert r.status_code == 200
        assert isinstance(r.json(), list)

    def test_issue_types(self, admin_token):
        r = requests.get(f"{BASE_URL}/api/issue-types",
                         headers=_auth(admin_token), timeout=10)
        assert r.status_code == 200
        assert len(r.json()) >= 1
