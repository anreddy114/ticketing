from dotenv import load_dotenv
from pathlib import Path

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")

import os
import logging
import uuid
import bcrypt
import jwt
from datetime import datetime, timezone, timedelta
from typing import List, Optional, Literal

from fastapi import FastAPI, APIRouter, HTTPException, Depends, Request, Response, Query
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
from pydantic import BaseModel, Field, EmailStr, ConfigDict
import httpx

# ---------- DB ----------
mongo_url = os.environ["MONGO_URL"]
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ["DB_NAME"]]

# ---------- App ----------
app = FastAPI(title="Ticketing System API")
api = APIRouter(prefix="/api")

JWT_ALGORITHM = "HS256"


def get_jwt_secret() -> str:
    return os.environ["JWT_SECRET"]


# ---------- Helpers ----------
def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def hash_password(pw: str) -> str:
    return bcrypt.hashpw(pw.encode(), bcrypt.gensalt()).decode()


def verify_password(pw: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(pw.encode(), hashed.encode())
    except Exception:
        return False


def create_access_token(user_id: str, email: str, role: str) -> str:
    payload = {
        "sub": user_id,
        "email": email,
        "role": role,
        "exp": datetime.now(timezone.utc) + timedelta(hours=12),
        "type": "access",
    }
    return jwt.encode(payload, get_jwt_secret(), algorithm=JWT_ALGORITHM)


def sanitize_user(u: dict) -> dict:
    return {
        "id": u["id"],
        "email": u["email"],
        "name": u["name"],
        "role": u["role"],
        "active": u.get("active", True),
        "online": bool(u.get("online", False)),
        "last_seen": u.get("last_seen"),
        "created_at": u.get("created_at"),
    }


async def get_current_user(request: Request) -> dict:
    token = request.cookies.get("access_token")
    if not token:
        auth = request.headers.get("Authorization", "")
        if auth.startswith("Bearer "):
            token = auth[7:]
    if not token:
        raise HTTPException(status_code=401, detail="Not authenticated")
    try:
        payload = jwt.decode(token, get_jwt_secret(), algorithms=[JWT_ALGORITHM])
        if payload.get("type") != "access":
            raise HTTPException(status_code=401, detail="Invalid token type")
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Invalid token")
    user = await db.users.find_one({"id": payload["sub"]}, {"_id": 0})
    if not user or not user.get("active", True):
        raise HTTPException(status_code=401, detail="User not found or inactive")
    return user


async def require_admin(user: dict = Depends(get_current_user)) -> dict:
    if user.get("role") != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    return user


# ---------- Models ----------
class LoginIn(BaseModel):
    email: EmailStr
    password: str


class RegisterIn(BaseModel):
    email: EmailStr
    password: str
    name: str
    role: Literal["admin", "agent"] = "agent"


class UserUpdate(BaseModel):
    name: Optional[str] = None
    role: Optional[Literal["admin", "agent"]] = None
    active: Optional[bool] = None
    password: Optional[str] = None


class IssueTypeIn(BaseModel):
    name: str
    description: Optional[str] = ""
    active: bool = True


class IssueTypeUpdate(BaseModel):
    name: Optional[str] = None
    description: Optional[str] = None
    active: Optional[bool] = None


class CustomerConfirmIn(BaseModel):
    mobile: str
    name: str


class TicketCreateIn(BaseModel):
    source: Literal["customer", "self"]
    # Customer source fields
    customer_mobile: Optional[str] = None
    customer_name: Optional[str] = None
    customer_email: Optional[str] = None
    customer_package: Optional[str] = None
    customer_expiry: Optional[str] = None
    customer_partner: Optional[str] = None
    customer_acc_id: Optional[int] = None
    # Common
    issue_type_id: str
    title: str
    description: str
    priority: Literal["low", "medium", "high", "urgent"] = "medium"
    assigned_to: Optional[str] = None  # user id; defaults to self


class TicketTransferIn(BaseModel):
    to_user_id: str
    note: Optional[str] = ""


class TicketCommentIn(BaseModel):
    message: str


class TicketStatusIn(BaseModel):
    status: Literal["open", "in_progress", "closed"]
    note: Optional[str] = ""
    resolution: Optional[str] = None  # required when closing


class PresenceIn(BaseModel):
    online: bool
    transfer_to: Optional[str] = None  # user id to reassign open tickets to when going offline


class SystemSettingsIn(BaseModel):
    offline_strategy: Optional[Literal["stay", "round_robin", "fallback", "manual_transfer"]] = None
    fallback_assignee_id: Optional[str] = None


class SipIncomingCallIn(BaseModel):
    caller_mobile: str
    call_id: Optional[str] = None
    did: Optional[str] = None
    agent_extension: Optional[str] = None
    agent_busy: Optional[bool] = True
    notes: Optional[str] = None


class IvrEventIn(BaseModel):
    caller_mobile: str
    event: str  # e.g. "dtmf", "menu_selected", "transfer", "hangup"
    payload: Optional[dict] = None
    call_id: Optional[str] = None


# ---------- Startup ----------
@app.on_event("startup")
async def startup():
    await db.users.create_index("email", unique=True)
    await db.users.create_index("id", unique=True)
    await db.tickets.create_index("id", unique=True)
    await db.tickets.create_index("ticket_number", unique=True)
    await db.issue_types.create_index("id", unique=True)
    await db.ticket_events.create_index("ticket_id")

    # Seed admin
    admin_email = os.environ.get("ADMIN_EMAIL", "admin@ticketing.com")
    admin_password = os.environ.get("ADMIN_PASSWORD", "admin123")
    existing = await db.users.find_one({"email": admin_email})
    if not existing:
        await db.users.insert_one({
            "id": str(uuid.uuid4()),
            "email": admin_email,
            "password_hash": hash_password(admin_password),
            "name": "Admin",
            "role": "admin",
            "active": True,
            "created_at": now_iso(),
        })
    elif not verify_password(admin_password, existing.get("password_hash", "")):
        await db.users.update_one(
            {"email": admin_email},
            {"$set": {"password_hash": hash_password(admin_password)}},
        )

    # Seed an agent for testing
    agent_email = "agent@ticketing.com"
    if not await db.users.find_one({"email": agent_email}):
        await db.users.insert_one({
            "id": str(uuid.uuid4()),
            "email": agent_email,
            "password_hash": hash_password("agent123"),
            "name": "Agent One",
            "role": "agent",
            "active": True,
            "created_at": now_iso(),
        })

    # Seed default issue types
    default_types = [
        ("Billing", "Billing or payment related issues"),
        ("Technical", "Technical product issues"),
        ("Service Request", "New service / change request"),
        ("Complaint", "Customer complaint"),
        ("General Inquiry", "General questions and information"),
    ]
    for n, d in default_types:
        if not await db.issue_types.find_one({"name": n}):
            await db.issue_types.insert_one({
                "id": str(uuid.uuid4()),
                "name": n,
                "description": d,
                "active": True,
                "created_at": now_iso(),
            })


@app.on_event("shutdown")
async def shutdown():
    client.close()


# ---------- Auth ----------
@api.post("/auth/login")
async def login(payload: LoginIn, response: Response):
    user = await db.users.find_one({"email": payload.email.lower()})
    if not user or not verify_password(payload.password, user.get("password_hash", "")):
        raise HTTPException(status_code=401, detail="Invalid email or password")
    if not user.get("active", True):
        raise HTTPException(status_code=403, detail="Account is inactive")
    token = create_access_token(user["id"], user["email"], user["role"])
    response.set_cookie(
        "access_token", token, httponly=True, secure=False, samesite="lax",
        max_age=12 * 3600, path="/",
    )
    # Start agent session
    session_id = str(uuid.uuid4())
    await db.agent_sessions.insert_one({
        "id": session_id,
        "user_id": user["id"],
        "user_name": user["name"],
        "login_at": now_iso(),
        "logout_at": None,
        "duration_sec": None,
    })
    await db.users.update_one(
        {"id": user["id"]},
        {"$set": {"current_session_id": session_id, "last_login_at": now_iso()}},
    )
    return {"token": token, "user": sanitize_user(user)}


@api.post("/auth/logout")
async def logout(response: Response, user: dict = Depends(get_current_user)):
    response.delete_cookie("access_token", path="/")
    # End agent session
    sid = user.get("current_session_id")
    if sid:
        sess = await db.agent_sessions.find_one({"id": sid})
        if sess and not sess.get("logout_at"):
            t0 = datetime.fromisoformat(sess["login_at"].replace("Z", "+00:00"))
            t1 = datetime.now(timezone.utc)
            await db.agent_sessions.update_one(
                {"id": sid},
                {"$set": {"logout_at": t1.isoformat(), "duration_sec": int((t1 - t0).total_seconds())}},
            )
    # Mark offline (but do NOT redistribute — they may log back in shortly)
    await db.users.update_one(
        {"id": user["id"]},
        {"$set": {"online": False, "current_session_id": None}},
    )
    return {"ok": True}


@api.get("/auth/me")
async def me(user: dict = Depends(get_current_user)):
    return sanitize_user(user)


@api.post("/auth/register")
async def register(payload: RegisterIn, _: dict = Depends(require_admin)):
    email = payload.email.lower()
    if await db.users.find_one({"email": email}):
        raise HTTPException(status_code=400, detail="Email already exists")
    user = {
        "id": str(uuid.uuid4()),
        "email": email,
        "password_hash": hash_password(payload.password),
        "name": payload.name,
        "role": payload.role,
        "active": True,
        "created_at": now_iso(),
    }
    await db.users.insert_one(user)
    return sanitize_user(user)


# ---------- Users (Admin) ----------
@api.get("/users")
async def list_users(user: dict = Depends(get_current_user)):
    users = await db.users.find({}, {"_id": 0, "password_hash": 0}).to_list(1000)
    return users


@api.patch("/users/{user_id}")
async def update_user(user_id: str, payload: UserUpdate, _: dict = Depends(require_admin)):
    update = {}
    if payload.name is not None:
        update["name"] = payload.name
    if payload.role is not None:
        update["role"] = payload.role
    if payload.active is not None:
        update["active"] = payload.active
    if payload.password:
        update["password_hash"] = hash_password(payload.password)
    if not update:
        raise HTTPException(status_code=400, detail="No changes")
    result = await db.users.update_one({"id": user_id}, {"$set": update})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    u = await db.users.find_one({"id": user_id}, {"_id": 0, "password_hash": 0})
    return u


@api.delete("/users/{user_id}")
async def delete_user(user_id: str, current: dict = Depends(require_admin)):
    if user_id == current["id"]:
        raise HTTPException(status_code=400, detail="Cannot delete yourself")
    result = await db.users.delete_one({"id": user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="User not found")
    return {"ok": True}


# ---------- Issue Types ----------
@api.get("/issue-types")
async def list_issue_types(_: dict = Depends(get_current_user)):
    return await db.issue_types.find({}, {"_id": 0}).to_list(1000)


@api.post("/issue-types")
async def create_issue_type(payload: IssueTypeIn, _: dict = Depends(require_admin)):
    if await db.issue_types.find_one({"name": payload.name}):
        raise HTTPException(status_code=400, detail="Issue type already exists")
    doc = {
        "id": str(uuid.uuid4()),
        "name": payload.name,
        "description": payload.description or "",
        "active": payload.active,
        "created_at": now_iso(),
    }
    await db.issue_types.insert_one(doc)
    doc.pop("_id", None)
    return doc


@api.patch("/issue-types/{type_id}")
async def update_issue_type(type_id: str, payload: IssueTypeUpdate, _: dict = Depends(require_admin)):
    update = {k: v for k, v in payload.model_dump(exclude_none=True).items()}
    if not update:
        raise HTTPException(status_code=400, detail="No changes")
    res = await db.issue_types.update_one({"id": type_id}, {"$set": update})
    if res.matched_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    return await db.issue_types.find_one({"id": type_id}, {"_id": 0})


@api.delete("/issue-types/{type_id}")
async def delete_issue_type(type_id: str, _: dict = Depends(require_admin)):
    res = await db.issue_types.delete_one({"id": type_id})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Not found")
    return {"ok": True}


# ---------- Customer lookup (SmartPlay portal API) ----------
@api.get("/customers/lookup")
async def lookup_customer(mobile: str = Query(..., min_length=4), _: dict = Depends(get_current_user)):
    base = os.environ.get("SMARTPLAY_API_URL", "").rstrip("/")
    token = os.environ.get("SMARTPLAY_API_TOKEN", "")
    if not base or not token:
        raise HTTPException(status_code=500, detail="Customer API not configured")
    url = f"{base}/api/smart-plays/mobile/{mobile.strip()}"
    try:
        async with httpx.AsyncClient(timeout=15) as cli:
            r = await cli.get(url, headers={"Authorization": f"Bearer {token}"})
    except httpx.HTTPError as e:
        raise HTTPException(status_code=502, detail=f"Customer API unreachable: {e}")

    if r.status_code != 200:
        raise HTTPException(status_code=404, detail="Customer not found")

    try:
        body = r.json()
    except Exception:
        raise HTTPException(status_code=502, detail="Invalid customer API response")

    if str(body.get("status")) != "200" or not body.get("results"):
        raise HTTPException(status_code=404, detail="Customer not found")

    res = body["results"]
    name = f"{(res.get('firstname') or '').strip()} {(res.get('lastname') or '').strip()}".strip() or "Unknown"
    return {
        "mobile": str(res.get("mobile") or mobile),
        "name": name,
        "email": res.get("email"),
        "acc_id": res.get("acc_id"),
        "expiry_date": res.get("expiry_date"),
        "partner": res.get("partner"),
        "partner_code": res.get("partner_code"),
        "package": res.get("package"),
    }


# ---------- WhatsApp (Meta Cloud API) ----------
def _format_msisdn(mobile: str) -> str:
    """Strip non-digits; if 10-digit local Indian number, prepend default country code."""
    digits = "".join(ch for ch in (mobile or "") if ch.isdigit())
    cc = os.environ.get("WHATSAPP_DEFAULT_COUNTRY_CODE", "91")
    if len(digits) == 10:
        digits = f"{cc}{digits}"
    return digits


async def send_whatsapp_message(mobile: str, message: str, ticket_id: str, kind: str, template_params: Optional[List[str]] = None):
    """Send a WhatsApp message via Meta Cloud API.

    For "created" and "closed" kinds we send an approved **template** (Meta requires
    a template for business-initiated messages). `message` is also stored locally
    so agents see what was conveyed even though Meta renders the approved template.

    Env vars:
      WHATSAPP_TEMPLATE_CREATED — template name for ticket-created notifications
      WHATSAPP_TEMPLATE_CLOSED  — template name for ticket-closed notifications
      WHATSAPP_TEMPLATE_LANG    — template language code (e.g. en_US)
    """
    version = os.environ.get("WHATSAPP_GRAPH_VERSION", "v21.0")
    phone_id = os.environ.get("WHATSAPP_PHONE_NUMBER_ID")
    token = os.environ.get("WHATSAPP_ACCESS_TOKEN")
    lang = os.environ.get("WHATSAPP_TEMPLATE_LANG", "en_US")
    tpl_map = {
        "created": os.environ.get("WHATSAPP_TEMPLATE_CREATED"),
        "closed": os.environ.get("WHATSAPP_TEMPLATE_CLOSED"),
    }
    template_name = tpl_map.get(kind)

    to = _format_msisdn(mobile)
    doc = {
        "id": str(uuid.uuid4()),
        "ticket_id": ticket_id,
        "mobile": to,
        "message": message,
        "kind": kind,
        "status": "pending",
        "provider": "meta_whatsapp_cloud",
        "provider_message_id": None,
        "provider_response": None,
        "template_name": template_name,
        "template_params": template_params or [],
        "error": None,
        "created_at": now_iso(),
    }

    if not phone_id or not token:
        doc["status"] = "skipped_not_configured"
        await db.whatsapp_messages.insert_one(doc)
        return doc

    url = f"https://graph.facebook.com/{version}/{phone_id}/messages"

    if template_name:
        components = []
        if template_params:
            components = [{
                "type": "body",
                "parameters": [{"type": "text", "text": str(p)} for p in template_params],
            }]
        payload = {
            "messaging_product": "whatsapp",
            "to": to,
            "type": "template",
            "template": {
                "name": template_name,
                "language": {"code": lang},
                **({"components": components} if components else {}),
            },
        }
    else:
        # Free-form text (only works inside an open 24h customer window)
        payload = {
            "messaging_product": "whatsapp",
            "recipient_type": "individual",
            "to": to,
            "type": "text",
            "text": {"preview_url": False, "body": message},
        }

    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    try:
        async with httpx.AsyncClient(timeout=20) as cli:
            r = await cli.post(url, json=payload, headers=headers)
        try:
            body = r.json()
        except Exception:
            body = {"raw": r.text}
        doc["provider_response"] = body
        if r.status_code in (200, 201) and isinstance(body, dict) and body.get("messages"):
            doc["status"] = "sent"
            doc["provider_message_id"] = body["messages"][0].get("id")
        else:
            doc["status"] = "failed"
            err = body.get("error") if isinstance(body, dict) else None
            if isinstance(err, dict):
                doc["error"] = f"{err.get('code', '')}: {err.get('message', '')} — {err.get('error_user_msg') or ''}".strip(" —")
            else:
                doc["error"] = f"HTTP {r.status_code}"
    except httpx.HTTPError as e:
        doc["status"] = "failed"
        doc["error"] = f"network: {e}"

    await db.whatsapp_messages.insert_one(doc)
    logging.info(f"[WhatsApp:{doc['status']}] -> {to} ({kind}/{template_name or 'text'})")
    return doc


@api.get("/whatsapp/messages")
async def list_whatsapp_messages(
    ticket_id: Optional[str] = None,
    _: dict = Depends(get_current_user),
):
    q = {}
    if ticket_id:
        q["ticket_id"] = ticket_id
    msgs = await db.whatsapp_messages.find(q, {"_id": 0}).sort("created_at", -1).to_list(200)
    return msgs


class WhatsappTestIn(BaseModel):
    mobile: str
    kind: Literal["created", "closed"] = "created"


@api.post("/whatsapp/test")
async def whatsapp_test(payload: WhatsappTestIn, _: dict = Depends(require_admin)):
    """Send a test WhatsApp message to verify Meta integration end-to-end."""
    if payload.kind == "created":
        tpl_params = ["Test User", "TKT-TEST", "Test Issue"]
    else:
        tpl_params = ["Test User", "TKT-TEST", "Test Issue", "Test Agent"]
    doc = await send_whatsapp_message(
        payload.mobile,
        f"[TEST] Ticket {payload.kind} notification — please ignore.",
        "test", payload.kind, template_params=tpl_params,
    )
    doc.pop("_id", None)
    return doc


# ---------- Tickets ----------
async def next_ticket_number() -> str:
    counter = await db.counters.find_one_and_update(
        {"_id": "ticket"},
        {"$inc": {"seq": 1}},
        upsert=True,
        return_document=True,
    )
    seq = counter["seq"] if counter else 1
    return f"TKT-{seq:05d}"


async def log_event(ticket_id: str, actor: dict, event_type: str, message: str = "", meta: Optional[dict] = None):
    doc = {
        "id": str(uuid.uuid4()),
        "ticket_id": ticket_id,
        "event_type": event_type,
        "message": message,
        "actor_id": actor["id"],
        "actor_name": actor["name"],
        "meta": meta or {},
        "created_at": now_iso(),
    }
    await db.ticket_events.insert_one(doc)


@api.post("/tickets")
async def create_ticket(payload: TicketCreateIn, user: dict = Depends(get_current_user)):
    issue = await db.issue_types.find_one({"id": payload.issue_type_id})
    if not issue:
        raise HTTPException(status_code=400, detail="Invalid issue type")

    assigned_to = payload.assigned_to or user["id"]
    assigned_user = await db.users.find_one({"id": assigned_to})
    if not assigned_user:
        raise HTTPException(status_code=400, detail="Assigned user not found")

    if payload.source == "customer":
        if not payload.customer_mobile or not payload.customer_name:
            raise HTTPException(status_code=400, detail="Customer mobile and name required")

    tnum = await next_ticket_number()
    ticket = {
        "id": str(uuid.uuid4()),
        "ticket_number": tnum,
        "source": payload.source,
        "customer_mobile": payload.customer_mobile if payload.source == "customer" else None,
        "customer_name": payload.customer_name if payload.source == "customer" else None,
        "customer_email": payload.customer_email if payload.source == "customer" else None,
        "customer_package": payload.customer_package if payload.source == "customer" else None,
        "customer_expiry": payload.customer_expiry if payload.source == "customer" else None,
        "customer_partner": payload.customer_partner if payload.source == "customer" else None,
        "customer_acc_id": payload.customer_acc_id if payload.source == "customer" else None,
        "issue_type_id": payload.issue_type_id,
        "issue_type_name": issue["name"],
        "title": payload.title,
        "description": payload.description,
        "priority": payload.priority,
        "status": "open",
        "created_by": user["id"],
        "created_by_name": user["name"],
        "assigned_to": assigned_to,
        "assigned_to_name": assigned_user["name"],
        "created_at": now_iso(),
        "updated_at": now_iso(),
        "closed_at": None,
    }
    await db.tickets.insert_one(ticket)
    await log_event(ticket["id"], user, "created", f"Ticket created and assigned to {assigned_user['name']}")

    # Send WhatsApp acknowledgement to customer
    if payload.source == "customer":
        msg = (
            f"Hi {payload.customer_name}, your ticket {tnum} has been created. "
            f"Issue: {issue['name']}. Our team will reach out shortly."
        )
        await send_whatsapp_message(
            payload.customer_mobile, msg, ticket["id"], "created",
            template_params=[payload.customer_name or "Customer", tnum, issue["name"]],
        )

    ticket.pop("_id", None)
    return ticket


@api.get("/tickets")
async def list_tickets(
    status: Optional[str] = None,
    assigned_to: Optional[str] = None,
    issue_type_id: Optional[str] = None,
    mine: Optional[bool] = False,
    search: Optional[str] = None,
    user: dict = Depends(get_current_user),
):
    q = {}
    if status:
        q["status"] = status
    if issue_type_id:
        q["issue_type_id"] = issue_type_id
    # Non-admins always see only their own tickets
    if user.get("role") != "admin":
        q["assigned_to"] = user["id"]
    elif mine:
        q["assigned_to"] = user["id"]
    elif assigned_to:
        q["assigned_to"] = assigned_to
    if search:
        q["$or"] = [
            {"ticket_number": {"$regex": search, "$options": "i"}},
            {"title": {"$regex": search, "$options": "i"}},
            {"customer_mobile": {"$regex": search, "$options": "i"}},
            {"customer_name": {"$regex": search, "$options": "i"}},
        ]
    tickets = await db.tickets.find(q, {"_id": 0}).sort("created_at", -1).to_list(500)
    return tickets


@api.get("/tickets/{ticket_id}")
async def get_ticket(ticket_id: str, _: dict = Depends(get_current_user)):
    t = await db.tickets.find_one({"id": ticket_id}, {"_id": 0})
    if not t:
        raise HTTPException(status_code=404, detail="Ticket not found")
    events = await db.ticket_events.find({"ticket_id": ticket_id}, {"_id": 0}).sort("created_at", 1).to_list(500)
    return {"ticket": t, "events": events}


@api.post("/tickets/{ticket_id}/transfer")
async def transfer_ticket(ticket_id: str, payload: TicketTransferIn, user: dict = Depends(get_current_user)):
    t = await db.tickets.find_one({"id": ticket_id})
    if not t:
        raise HTTPException(status_code=404, detail="Ticket not found")
    if t["status"] == "closed":
        raise HTTPException(status_code=400, detail="Cannot transfer a closed ticket")
    to_user = await db.users.find_one({"id": payload.to_user_id})
    if not to_user:
        raise HTTPException(status_code=400, detail="Target user not found")
    if to_user["id"] == t["assigned_to"]:
        raise HTTPException(status_code=400, detail="Ticket already assigned to this user")
    prev_name = t.get("assigned_to_name", "")
    await db.tickets.update_one(
        {"id": ticket_id},
        {"$set": {
            "assigned_to": to_user["id"],
            "assigned_to_name": to_user["name"],
            "updated_at": now_iso(),
        }},
    )
    await log_event(
        ticket_id, user, "transferred",
        f"Transferred from {prev_name} to {to_user['name']}." + (f" Note: {payload.note}" if payload.note else ""),
        meta={"from": t["assigned_to"], "to": to_user["id"]},
    )
    return await db.tickets.find_one({"id": ticket_id}, {"_id": 0})


@api.post("/tickets/{ticket_id}/comment")
async def add_comment(ticket_id: str, payload: TicketCommentIn, user: dict = Depends(get_current_user)):
    t = await db.tickets.find_one({"id": ticket_id})
    if not t:
        raise HTTPException(status_code=404, detail="Ticket not found")
    await log_event(ticket_id, user, "comment", payload.message)
    await db.tickets.update_one({"id": ticket_id}, {"$set": {"updated_at": now_iso()}})
    return {"ok": True}


@api.post("/tickets/{ticket_id}/status")
async def change_status(ticket_id: str, payload: TicketStatusIn, user: dict = Depends(get_current_user)):
    t = await db.tickets.find_one({"id": ticket_id})
    if not t:
        raise HTTPException(status_code=404, detail="Ticket not found")
    if t["status"] == payload.status:
        raise HTTPException(status_code=400, detail="Ticket already in this status")
    if payload.status == "closed" and not (payload.resolution or "").strip():
        raise HTTPException(status_code=400, detail="Resolution message is required to close a ticket")
    update = {"status": payload.status, "updated_at": now_iso()}
    if payload.status == "closed":
        update["closed_at"] = now_iso()
        update["resolution"] = payload.resolution.strip()
        update["closed_by"] = user["id"]
        update["closed_by_name"] = user["name"]
    await db.tickets.update_one({"id": ticket_id}, {"$set": update})
    msg = f"Status changed from {t['status']} to {payload.status}"
    if payload.status == "closed":
        msg += f". Resolution: {payload.resolution.strip()}"
    if payload.note:
        msg += f". Note: {payload.note}"
    await log_event(
        ticket_id, user, "status_change", msg,
        meta={"from": t["status"], "to": payload.status, "resolution": payload.resolution if payload.status == "closed" else None},
    )

    # Send WhatsApp on close
    if payload.status == "closed" and t.get("source") == "customer" and t.get("customer_mobile"):
        wa_msg = (
            f"Hi {t.get('customer_name', 'Customer')}, your ticket {t['ticket_number']} "
            f"has been closed. Thank you for reaching out."
        )
        await send_whatsapp_message(
            t["customer_mobile"], wa_msg, ticket_id, "closed",
            template_params=[
                t.get("customer_name") or "Customer",
                t["ticket_number"],
                t.get("issue_type_name") or "—",
                t.get("assigned_to_name") or "Support Team",
            ],
        )

    return await db.tickets.find_one({"id": ticket_id}, {"_id": 0})


# ---------- Reports ----------
@api.get("/reports/summary")
async def reports_summary(user: dict = Depends(get_current_user)):
    match: dict = {}
    if user.get("role") != "admin":
        match["assigned_to"] = user["id"]
    total = await db.tickets.count_documents(match)
    open_c = await db.tickets.count_documents({**match, "status": "open"})
    in_prog = await db.tickets.count_documents({**match, "status": "in_progress"})
    closed = await db.tickets.count_documents({**match, "status": "closed"})

    base_pipeline = [{"$match": match}] if match else []

    # By issue type
    pipe_type = base_pipeline + [
        {"$group": {"_id": "$issue_type_name", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}},
    ]
    by_type = await db.tickets.aggregate(pipe_type).to_list(100)
    by_type = [{"name": x["_id"] or "Unknown", "count": x["count"]} for x in by_type]

    # By assignee
    pipe_assignee = base_pipeline + [
        {"$group": {"_id": "$assigned_to_name", "count": {"$sum": 1}}},
        {"$sort": {"count": -1}},
    ]
    by_assignee = await db.tickets.aggregate(pipe_assignee).to_list(100)
    by_assignee = [{"name": x["_id"] or "Unknown", "count": x["count"]} for x in by_assignee]

    # By priority
    pipe_priority = base_pipeline + [
        {"$group": {"_id": "$priority", "count": {"$sum": 1}}},
    ]
    by_priority = await db.tickets.aggregate(pipe_priority).to_list(100)
    by_priority = [{"name": x["_id"] or "Unknown", "count": x["count"]} for x in by_priority]

    return {
        "totals": {"total": total, "open": open_c, "in_progress": in_prog, "closed": closed},
        "by_issue_type": by_type,
        "by_assignee": by_assignee,
        "by_priority": by_priority,
    }


@api.get("/reports/leaderboard")
async def reports_leaderboard(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    sla_hours: int = 24,
    _: dict = Depends(require_admin),
):
    """Per-agent productivity leaderboard.
    Computes: assigned/open/in_progress/closed counts, avg resolution time,
    total logged hours, tickets-closed-per-hour, and SLA breaches.
    """
    dt_from = _parse_dt(date_from)
    dt_to = _parse_dt(date_to)
    ticket_match: dict = {}
    if dt_from or dt_to:
        rng: dict = {}
        if dt_from:
            rng["$gte"] = dt_from.isoformat()
        if dt_to:
            rng["$lte"] = dt_to.isoformat()
        ticket_match["created_at"] = rng

    users = await db.users.find({"active": True}, {"_id": 0, "password_hash": 0}).to_list(1000)
    now_t = datetime.now(timezone.utc)
    sla_threshold_sec = sla_hours * 3600
    rows = []
    for u in users:
        uid = u["id"]
        umatch = {**ticket_match, "assigned_to": uid}
        all_t = await db.tickets.find(umatch, {"_id": 0}).to_list(5000)
        total = len(all_t)
        open_c = sum(1 for t in all_t if t["status"] == "open")
        in_prog = sum(1 for t in all_t if t["status"] == "in_progress")
        closed_list = [t for t in all_t if t["status"] == "closed"]
        closed = len(closed_list)

        # Avg resolution (closed only)
        res_secs = []
        for t in closed_list:
            try:
                t0 = datetime.fromisoformat(t["created_at"].replace("Z", "+00:00"))
                t1 = datetime.fromisoformat(t["closed_at"].replace("Z", "+00:00"))
                res_secs.append((t1 - t0).total_seconds())
            except Exception:
                pass
        avg_res = (sum(res_secs) / len(res_secs)) if res_secs else None

        # Logged time from sessions
        sess_match: dict = {"user_id": uid}
        if dt_from or dt_to:
            rng: dict = {}
            if dt_from:
                rng["$gte"] = dt_from.isoformat()
            if dt_to:
                rng["$lte"] = dt_to.isoformat()
            sess_match["login_at"] = rng
        sessions = await db.agent_sessions.find(sess_match, {"_id": 0}).to_list(2000)
        logged = 0
        for s in sessions:
            if s.get("duration_sec") is not None:
                logged += s["duration_sec"]
            else:
                try:
                    t0 = datetime.fromisoformat(s["login_at"].replace("Z", "+00:00"))
                    logged += int((now_t - t0).total_seconds())
                except Exception:
                    pass

        # SLA breaches: open or in_progress past threshold; closed but resolution > threshold
        breaches = 0
        for t in all_t:
            try:
                t0 = datetime.fromisoformat(t["created_at"].replace("Z", "+00:00"))
            except Exception:
                continue
            if t["status"] in ("open", "in_progress"):
                if (now_t - t0).total_seconds() > sla_threshold_sec:
                    breaches += 1
            elif t["status"] == "closed" and t.get("closed_at"):
                try:
                    t1 = datetime.fromisoformat(t["closed_at"].replace("Z", "+00:00"))
                    if (t1 - t0).total_seconds() > sla_threshold_sec:
                        breaches += 1
                except Exception:
                    pass

        hours = logged / 3600 if logged else 0
        per_hour = (closed / hours) if hours > 0 else None
        rows.append({
            "user_id": uid,
            "name": u["name"],
            "role": u["role"],
            "online": bool(u.get("online")),
            "total": total,
            "open": open_c,
            "in_progress": in_prog,
            "closed": closed,
            "avg_resolution_sec": int(avg_res) if avg_res is not None else None,
            "logged_sec": logged,
            "closed_per_hour": round(per_hour, 2) if per_hour is not None else None,
            "sla_breaches": breaches,
        })

    # rank by closed desc, then closed_per_hour desc, then total desc
    rows.sort(key=lambda r: (-r["closed"], -(r["closed_per_hour"] or 0), -r["total"]))
    return {"sla_hours": sla_hours, "rows": rows}


@api.get("/")
async def root():
    return {"service": "Ticketing System API", "status": "ok"}


# ====================================================================
# Agent Presence (online/offline + heartbeat)
# ====================================================================

HEARTBEAT_TIMEOUT_SEC = int(os.environ.get("AGENT_HEARTBEAT_TIMEOUT_SEC", "120"))


def _seconds_since(iso: Optional[str]) -> float:
    if not iso:
        return 1e9
    try:
        t = datetime.fromisoformat(iso.replace("Z", "+00:00"))
        return (datetime.now(timezone.utc) - t).total_seconds()
    except Exception:
        return 1e9


async def _is_user_online(u: dict) -> bool:
    if not u.get("online"):
        return False
    return _seconds_since(u.get("last_seen")) < HEARTBEAT_TIMEOUT_SEC


async def _online_agents(exclude_id: Optional[str] = None) -> List[dict]:
    users = await db.users.find({"active": True}, {"_id": 0, "password_hash": 0}).to_list(1000)
    out = []
    for u in users:
        if exclude_id and u["id"] == exclude_id:
            continue
        if await _is_user_online(u):
            out.append(u)
    return out


async def _round_robin_pick(exclude_id: Optional[str] = None) -> Optional[dict]:
    agents = await _online_agents(exclude_id=exclude_id)
    if not agents:
        return None
    counter = await db.counters.find_one_and_update(
        {"_id": "rr"}, {"$inc": {"seq": 1}}, upsert=True, return_document=True,
    )
    idx = (counter["seq"] - 1) % len(agents)
    return agents[idx]


async def _get_settings() -> dict:
    s = await db.system_settings.find_one({"_id": "global"}, {"_id": 0}) or {}
    s.setdefault("offline_strategy", os.environ.get("OFFLINE_STRATEGY", "round_robin"))
    s.setdefault("fallback_assignee_id", None)
    return s


async def _resolve_fallback_user() -> Optional[dict]:
    settings = await _get_settings()
    fid = settings.get("fallback_assignee_id")
    if fid:
        u = await db.users.find_one({"id": fid, "active": True})
        if u:
            return u
    fallback_email = os.environ.get("FALLBACK_ASSIGNEE_EMAIL", "admin@ticketing.com")
    return await db.users.find_one({"email": fallback_email, "active": True})


async def _redistribute_open_tickets(from_user: dict, strategy: str, transfer_to_id: Optional[str] = None) -> int:
    """Move open tickets owned by `from_user` to another agent based on strategy."""
    tickets = await db.tickets.find(
        {"assigned_to": from_user["id"], "status": {"$in": ["open", "in_progress"]}}
    ).to_list(2000)
    if not tickets:
        return 0

    target: Optional[dict] = None
    if strategy == "manual_transfer" and transfer_to_id:
        target = await db.users.find_one({"id": transfer_to_id, "active": True})
    elif strategy == "fallback":
        target = await _resolve_fallback_user()
        if target and target["id"] == from_user["id"]:
            target = None

    moved = 0
    for t in tickets:
        chosen = target
        if strategy == "round_robin" and not chosen:
            chosen = await _round_robin_pick(exclude_id=from_user["id"])
        if not chosen:
            continue  # leave assigned (no available target)
        if chosen["id"] == t["assigned_to"]:
            continue
        await db.tickets.update_one(
            {"id": t["id"]},
            {"$set": {
                "assigned_to": chosen["id"],
                "assigned_to_name": chosen["name"],
                "updated_at": now_iso(),
            }},
        )
        await log_event(
            t["id"], from_user, "transferred",
            f"Auto-transferred from {from_user['name']} to {chosen['name']} (agent went offline / {strategy}).",
            meta={"from": from_user["id"], "to": chosen["id"], "reason": "presence_offline", "strategy": strategy},
        )
        moved += 1
    return moved


@api.get("/agents")
async def list_agents(_: dict = Depends(get_current_user)):
    """List agents with computed online status."""
    users = await db.users.find({"active": True}, {"_id": 0, "password_hash": 0}).to_list(1000)
    out = []
    for u in users:
        u["online"] = await _is_user_online(u)
        out.append(u)
    return out


@api.post("/agents/heartbeat")
async def heartbeat(user: dict = Depends(get_current_user)):
    """Frontend pings this every ~60s. Keeps `last_seen` fresh."""
    await db.users.update_one(
        {"id": user["id"]},
        {"$set": {"last_seen": now_iso()}, "$setOnInsert": {}},
    )
    return {"ok": True, "at": now_iso()}


@api.post("/agents/presence")
async def set_presence(payload: PresenceIn, user: dict = Depends(get_current_user)):
    """Toggle agent online/offline.

    When going offline, redistribute open tickets per system offline_strategy.
    The frontend may pass `transfer_to` (user id) for the 'manual_transfer' strategy.
    Also logs an admin notification.
    """
    moved = 0
    previously_online = bool(user.get("online"))
    if not payload.online and previously_online:
        settings = await _get_settings()
        strategy = settings["offline_strategy"]
        if strategy == "manual_transfer" and not payload.transfer_to:
            raise HTTPException(
                status_code=400,
                detail="Pick a colleague to receive your open tickets before going offline.",
            )
        moved = await _redistribute_open_tickets(user, strategy, payload.transfer_to)

    await db.users.update_one(
        {"id": user["id"]},
        {"$set": {"online": payload.online, "last_seen": now_iso()}},
    )

    # Drop an admin notification (only when state actually changes)
    if previously_online != payload.online:
        notif = {
            "id": str(uuid.uuid4()),
            "kind": "agent_online" if payload.online else "agent_offline",
            "actor_id": user["id"],
            "actor_name": user["name"],
            "message": f"{user['name']} is now {'online' if payload.online else 'offline'}",
            "tickets_reassigned": moved,
            "for_admin": True,
            "read_by": [],
            "created_at": now_iso(),
        }
        await db.notifications.insert_one(notif)

    return {"ok": True, "online": payload.online, "tickets_reassigned": moved}


# ====================================================================
# Notifications (admin)
# ====================================================================

@api.get("/notifications")
async def list_notifications(user: dict = Depends(get_current_user), unread_only: bool = False):
    if user.get("role") != "admin":
        return []
    q: dict = {"for_admin": True}
    if unread_only:
        q["read_by"] = {"$nin": [user["id"]]}
    notifs = await db.notifications.find(q, {"_id": 0}).sort("created_at", -1).to_list(100)
    return notifs


@api.post("/notifications/mark-read")
async def mark_notifications_read(user: dict = Depends(require_admin)):
    await db.notifications.update_many(
        {"for_admin": True, "read_by": {"$nin": [user["id"]]}},
        {"$addToSet": {"read_by": user["id"]}},
    )
    return {"ok": True}


# ====================================================================
# Agent sessions (login/logout audit)
# ====================================================================

@api.get("/agents/sessions")
async def list_sessions(
    user_id: Optional[str] = None,
    user: dict = Depends(get_current_user),
):
    """Admins see all sessions; agents see only their own."""
    q: dict = {}
    if user.get("role") != "admin":
        q["user_id"] = user["id"]
    elif user_id:
        q["user_id"] = user_id
    sessions = await db.agent_sessions.find(q, {"_id": 0}).sort("login_at", -1).to_list(500)
    # compute live duration for active sessions
    now_t = datetime.now(timezone.utc)
    for s in sessions:
        if not s.get("logout_at"):
            try:
                t0 = datetime.fromisoformat(s["login_at"].replace("Z", "+00:00"))
                s["duration_sec"] = int((now_t - t0).total_seconds())
                s["active"] = True
            except Exception:
                s["active"] = True
        else:
            s["active"] = False
    return sessions


# ====================================================================
# Public website API — for customer-facing site to create tickets
# ====================================================================

def _require_public_key(request: Request):
    expected = os.environ.get("PUBLIC_API_KEY", "")
    if not expected:
        return  # not configured → allow (dev mode)
    got = request.headers.get("X-Public-Api-Key") or request.query_params.get("api_key")
    if got != expected:
        raise HTTPException(status_code=401, detail="Invalid public API key")


class PublicTicketIn(BaseModel):
    customer_mobile: str
    customer_name: Optional[str] = None
    issue_type_id: Optional[str] = None
    issue_type_name: Optional[str] = None
    title: str
    description: str
    priority: Literal["low", "medium", "high", "urgent"] = "medium"


@api.get("/public/customers/lookup")
async def public_lookup_customer(mobile: str = Query(..., min_length=4), request: Request = None):
    """Same as internal lookup but auth via public API key. Use this from your website."""
    _require_public_key(request)
    cust = await _smartplay_lookup_silent(mobile)
    if not cust:
        raise HTTPException(status_code=404, detail="Customer not found")
    return cust


@api.post("/public/tickets")
async def public_create_ticket(payload: PublicTicketIn, request: Request):
    """Public endpoint your website calls to create a ticket on behalf of a customer."""
    _require_public_key(request)

    # Resolve issue type
    issue = None
    if payload.issue_type_id:
        issue = await db.issue_types.find_one({"id": payload.issue_type_id, "active": True})
    if not issue and payload.issue_type_name:
        issue = await db.issue_types.find_one({"name": payload.issue_type_name, "active": True})
    if not issue:
        issue = (
            await db.issue_types.find_one({"name": "General Inquiry", "active": True})
            or await db.issue_types.find_one({"active": True})
        )
    if not issue:
        raise HTTPException(status_code=500, detail="No issue types configured")

    # Enrich with SmartPlay if name missing
    customer = await _smartplay_lookup_silent(payload.customer_mobile) or {}
    customer_name = (payload.customer_name or customer.get("name") or "Customer").strip()

    # Auto-assign to an online agent (round-robin), else fallback
    assignee = await _round_robin_pick()
    if not assignee:
        assignee = await _resolve_fallback_user()
    if not assignee:
        raise HTTPException(status_code=500, detail="No assignee available")

    tnum = await next_ticket_number()
    ticket = {
        "id": str(uuid.uuid4()),
        "ticket_number": tnum,
        "source": "customer",
        "customer_mobile": payload.customer_mobile.strip(),
        "customer_name": customer_name,
        "customer_email": customer.get("email"),
        "customer_package": customer.get("package"),
        "customer_expiry": customer.get("expiry_date"),
        "customer_partner": customer.get("partner"),
        "customer_acc_id": customer.get("acc_id"),
        "issue_type_id": issue["id"],
        "issue_type_name": issue["name"],
        "title": payload.title.strip(),
        "description": payload.description.strip(),
        "priority": payload.priority,
        "status": "open",
        "created_by": "public_website",
        "created_by_name": "Website",
        "assigned_to": assignee["id"],
        "assigned_to_name": assignee["name"],
        "created_at": now_iso(),
        "updated_at": now_iso(),
        "closed_at": None,
        "origin_channel": "public_website",
    }
    await db.tickets.insert_one(ticket)

    actor = {"id": "public_website", "name": "Website"}
    await log_event(
        ticket["id"], actor, "created",
        f"Ticket created from customer website. Auto-assigned to {assignee['name']}.",
    )

    await send_whatsapp_message(
        ticket["customer_mobile"],
        f"Hi {customer_name}, your ticket {tnum} has been created. Issue: {issue['name']}. Our team will reach out shortly.",
        ticket["id"], "created",
        template_params=[customer_name, tnum, issue["name"]],
    )

    ticket.pop("_id", None)
    return ticket


@api.get("/public/issue-types")
async def public_issue_types(request: Request):
    """Public — list active issue types so your website can render a dropdown."""
    _require_public_key(request)
    items = await db.issue_types.find({"active": True}, {"_id": 0, "description": 0}).to_list(100)
    return items


# ====================================================================
# System Settings (admin)
# ====================================================================

@api.get("/settings")
async def get_settings(_: dict = Depends(get_current_user)):
    s = await _get_settings()
    return {
        "offline_strategy": s["offline_strategy"],
        "fallback_assignee_id": s.get("fallback_assignee_id"),
        "sip_webhook_url": "/api/sip/incoming-call",
        "ivr_webhook_url": "/api/ivr/event",
        "heartbeat_timeout_sec": HEARTBEAT_TIMEOUT_SEC,
    }


@api.patch("/settings")
async def update_settings(payload: SystemSettingsIn, _: dict = Depends(require_admin)):
    update = {k: v for k, v in payload.model_dump(exclude_none=True).items()}
    if not update:
        raise HTTPException(status_code=400, detail="No changes")
    await db.system_settings.update_one({"_id": "global"}, {"$set": update}, upsert=True)
    return await _get_settings()


# ====================================================================
# SIP (Asterisk) — incoming call webhook
# ====================================================================

def _require_sip_secret(request: Request):
    expected = os.environ.get("SIP_WEBHOOK_SECRET", "")
    if not expected:
        return  # no secret configured → allow (dev mode)
    got = request.headers.get("X-Sip-Secret") or request.query_params.get("secret")
    if got != expected:
        raise HTTPException(status_code=401, detail="Invalid SIP webhook secret")


async def _smartplay_lookup_silent(mobile: str) -> Optional[dict]:
    base = os.environ.get("SMARTPLAY_API_URL", "").rstrip("/")
    token = os.environ.get("SMARTPLAY_API_TOKEN", "")
    if not base or not token:
        return None
    try:
        async with httpx.AsyncClient(timeout=10) as cli:
            r = await cli.get(
                f"{base}/api/smart-plays/mobile/{mobile.strip()}",
                headers={"Authorization": f"Bearer {token}"},
            )
        if r.status_code != 200:
            return None
        body = r.json()
        if str(body.get("status")) != "200" or not body.get("results"):
            return None
        res = body["results"]
        name = f"{(res.get('firstname') or '').strip()} {(res.get('lastname') or '').strip()}".strip() or None
        return {
            "mobile": str(res.get("mobile") or mobile),
            "name": name,
            "email": res.get("email"),
            "acc_id": res.get("acc_id"),
            "expiry_date": res.get("expiry_date"),
            "partner": res.get("partner"),
            "partner_code": res.get("partner_code"),
            "package": res.get("package"),
        }
    except Exception:
        return None


@api.post("/sip/incoming-call")
async def sip_incoming_call(payload: SipIncomingCallIn, request: Request):
    """Webhook for Asterisk dialplan/AGI.

    Asterisk should POST when an inbound call cannot be answered (no agent
    available / agent busy / queue timeout). We create a ticket automatically.

    Headers: `X-Sip-Secret: <SIP_WEBHOOK_SECRET>` (set in backend .env)
    """
    _require_sip_secret(request)

    # Try to enrich with SmartPlay customer data
    customer = await _smartplay_lookup_silent(payload.caller_mobile)
    customer_name = (customer or {}).get("name") or "Unknown Caller"

    # Pick issue type (telephony / general inquiry)
    issue = (
        await db.issue_types.find_one({"name": "General Inquiry"})
        or await db.issue_types.find_one({"active": True})
    )
    if not issue:
        raise HTTPException(status_code=500, detail="No issue types configured")

    # Pick assignee: round-robin among online agents, else fallback
    assignee = await _round_robin_pick()
    if not assignee:
        assignee = await _resolve_fallback_user()
    if not assignee:
        raise HTTPException(status_code=500, detail="No assignee available")

    tnum = await next_ticket_number()
    description_lines = [
        "Auto-created from missed/queued SIP call.",
        f"Call ID: {payload.call_id or '—'}",
        f"DID: {payload.did or '—'}",
        f"Agent busy: {payload.agent_busy}",
    ]
    if payload.notes:
        description_lines.append(f"Notes: {payload.notes}")

    ticket = {
        "id": str(uuid.uuid4()),
        "ticket_number": tnum,
        "source": "customer",
        "customer_mobile": (customer or {}).get("mobile") or payload.caller_mobile,
        "customer_name": customer_name,
        "customer_email": (customer or {}).get("email"),
        "customer_package": (customer or {}).get("package"),
        "customer_expiry": (customer or {}).get("expiry_date"),
        "customer_partner": (customer or {}).get("partner"),
        "customer_acc_id": (customer or {}).get("acc_id"),
        "issue_type_id": issue["id"],
        "issue_type_name": issue["name"],
        "title": f"Inbound call from {customer_name} ({payload.caller_mobile})",
        "description": "\n".join(description_lines),
        "priority": "high" if payload.agent_busy else "medium",
        "status": "open",
        "created_by": "system_sip",
        "created_by_name": "SIP Webhook",
        "assigned_to": assignee["id"],
        "assigned_to_name": assignee["name"],
        "created_at": now_iso(),
        "updated_at": now_iso(),
        "closed_at": None,
        "origin_channel": "sip_inbound",
        "origin_call_id": payload.call_id,
    }
    await db.tickets.insert_one(ticket)

    actor = {"id": "system_sip", "name": "SIP Webhook"}
    await log_event(
        ticket["id"], actor, "created",
        f"Auto-created from inbound SIP call. Assigned to {assignee['name']}.",
        meta={"call_id": payload.call_id, "did": payload.did, "agent_busy": payload.agent_busy},
    )

    # Send WhatsApp acknowledgement (template)
    await send_whatsapp_message(
        ticket["customer_mobile"],
        f"Hi {customer_name}, your call has been logged as ticket {tnum}. Our team will reach out shortly.",
        ticket["id"], "created",
        template_params=[customer_name, tnum, issue["name"]],
    )

    ticket.pop("_id", None)
    return ticket


# ====================================================================
# IVR webhook
# ====================================================================

@api.post("/ivr/event")
async def ivr_event(payload: IvrEventIn, request: Request):
    """Generic webhook for IVR systems to log customer interactions.
    Doesn't necessarily create a ticket — just records the touchpoint.
    """
    _require_sip_secret(request)
    doc = {
        "id": str(uuid.uuid4()),
        "caller_mobile": payload.caller_mobile,
        "event": payload.event,
        "payload": payload.payload or {},
        "call_id": payload.call_id,
        "created_at": now_iso(),
    }
    await db.ivr_events.insert_one(doc)
    doc.pop("_id", None)
    return {"ok": True, "event": doc}


@api.get("/ivr/agent-availability")
async def ivr_agent_availability(request: Request):
    """IVR can call this to decide whether to queue or send to voicemail/ticket."""
    _require_sip_secret(request)
    agents = await _online_agents()
    return {"online_count": len(agents), "available": len(agents) > 0}


# ====================================================================
# Reports — Excel export
# ====================================================================

def _parse_dt(s: Optional[str]) -> Optional[datetime]:
    if not s:
        return None
    try:
        return datetime.fromisoformat(s.replace("Z", "+00:00"))
    except Exception:
        return None


@api.get("/reports/tickets.xlsx")
async def export_tickets_xlsx(
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    status: Optional[str] = None,
    assigned_to: Optional[str] = None,
    issue_type_id: Optional[str] = None,
    user: dict = Depends(get_current_user),
):
    """Download tickets as Excel. Agents see only their own tickets unless admin."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    from fastapi.responses import StreamingResponse

    q: dict = {}
    if status:
        q["status"] = status
    if issue_type_id:
        q["issue_type_id"] = issue_type_id
    if assigned_to:
        q["assigned_to"] = assigned_to
    if user.get("role") != "admin":
        q["assigned_to"] = user["id"]  # agents export only theirs

    dt_from = _parse_dt(date_from)
    dt_to = _parse_dt(date_to)
    if dt_from or dt_to:
        rng: dict = {}
        if dt_from:
            rng["$gte"] = dt_from.isoformat()
        if dt_to:
            rng["$lte"] = dt_to.isoformat()
        q["created_at"] = rng

    tickets = await db.tickets.find(q, {"_id": 0}).sort("created_at", -1).to_list(20000)

    wb = Workbook()
    ws = wb.active
    ws.title = "Tickets"
    headers = [
        "Ticket #", "Title", "Status", "Priority", "Issue Type", "Source",
        "Customer Name", "Customer Mobile", "Package", "Expiry", "Partner",
        "Assigned To", "Created By", "Created At", "Closed At",
    ]
    ws.append(headers)
    hdr_fill = PatternFill(start_color="0A0A0A", end_color="0A0A0A", fill_type="solid")
    hdr_font = Font(bold=True, color="FFFFFF")
    for col, _ in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(vertical="center")

    for t in tickets:
        ws.append([
            t.get("ticket_number"),
            t.get("title"),
            t.get("status"),
            t.get("priority"),
            t.get("issue_type_name"),
            t.get("source"),
            t.get("customer_name") or "",
            t.get("customer_mobile") or "",
            t.get("customer_package") or "",
            t.get("customer_expiry") or "",
            t.get("customer_partner") or "",
            t.get("assigned_to_name") or "",
            t.get("created_by_name") or "",
            t.get("created_at") or "",
            t.get("closed_at") or "",
        ])

    # autosize columns (approx)
    for col_idx, header in enumerate(headers, start=1):
        max_len = len(str(header))
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx, min_row=2):
            for cell in row:
                v = "" if cell.value is None else str(cell.value)
                if len(v) > max_len:
                    max_len = len(v)
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = min(max_len + 2, 40)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"tickets_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


app.include_router(api)
app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get("CORS_ORIGINS", "*").split(","),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(name)s - %(levelname)s - %(message)s")
